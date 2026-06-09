"""
Actualiza el Excel de casos de prueba con los resultados de ejecución de Behave.

Mapeo: tags @U-X.Y / @E-X.Y / @I-X.Y  →  columna "Id Caso de Prueba" en hoja Escenarios.
Columnas actualizadas:
  J (10) - Resultado obtenido
  L (12) - Ciclo 1  →  "Pasó" | "Falló"
"""

import os
import shutil
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    raise ImportError("Instala openpyxl: pip install openpyxl")

EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), "Escenarios", "Ejemplo_Casos de prueba.xlsx"
)

COL_ID_CASO   = 6   # F  - Id Caso de Prueba
COL_RESULTADO = 10  # J  - Resultado obtenido
COL_CICLO1    = 12  # L  - Ciclo 1

FILL_PASS = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_FAIL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FONT_PASS = Font(color="276221", bold=True)
FONT_FAIL = Font(color="9C0006", bold=True)


def _tag_matches(excel_id: str, tag: str) -> bool:
    """True si el ID del Excel empieza con el tag (sin @)."""
    return excel_id.startswith(tag)


def update_excel(results: dict, excel_path: str = EXCEL_PATH) -> int:
    """
    Actualiza el Excel con los resultados.

    results = {
        'U-13.1': {
            'status': 'passed' | 'failed',
            'scenario_name': str,
            'error_msg': str | None
        }, ...
    }

    Retorna el número de filas actualizadas.
    """
    if not os.path.exists(excel_path):
        print(f"[EXCEL] Archivo no encontrado: {excel_path}")
        return 0

    # Backup antes de modificar
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = excel_path.replace(".xlsx", f"_backup_{ts}.xlsx")
    shutil.copy2(excel_path, backup)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Escenarios"]

    updated = 0
    # Construir lookup: prefijo → datos resultado
    # Un escenario puede tener varios tags; results tiene un entry por tag
    for row_idx in range(3, ws.max_row + 1):
        cell_id = ws.cell(row=row_idx, column=COL_ID_CASO)
        if not cell_id.value:
            continue

        excel_id = str(cell_id.value).strip()

        matched_tag = None
        for tag in results:
            if _tag_matches(excel_id, tag):
                matched_tag = tag
                break

        if matched_tag is None:
            continue

        data = results[matched_tag]
        passed = data["status"] == "passed"

        # Ciclo 1
        cell_ciclo = ws.cell(row=row_idx, column=COL_CICLO1)
        cell_ciclo.value = "Pasó" if passed else "Falló"
        cell_ciclo.fill = FILL_PASS if passed else FILL_FAIL
        cell_ciclo.font = FONT_PASS if passed else FONT_FAIL

        # Resultado obtenido (solo actualiza si está vacío O si hay error)
        cell_res = ws.cell(row=row_idx, column=COL_RESULTADO)
        if passed:
            if not cell_res.value:
                cell_res.value = "Escenario automatizado ejecutado correctamente."
        else:
            err = data.get("error_msg") or "Escenario falló. Ver reporte HTML."
            # Truncar a 500 chars para no desbordar la celda
            cell_res.value = err[:500]

        updated += 1

    wb.save(excel_path)
    print(f"[EXCEL] {updated} fila(s) actualizadas en {os.path.basename(excel_path)}")
    print(f"[EXCEL] Backup guardado: {os.path.basename(backup)}")
    return updated
